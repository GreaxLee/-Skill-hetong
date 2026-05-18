#!/usr/bin/env python3
"""
generate_contract.py — Automatically generate purchase contracts from an order plan.

Usage:
    python generate_contract.py 2026-06-01

Configuration:
    Edit config.json (same folder as this script) to change file paths or add/modify factories.
    Price list columns: SKU | 工厂 | 出厂价 | 含税价 | 产品名称 | 参数
"""

import sys
import os
import re
import json
import shutil
import zipfile
import datetime
from copy import copy
from typing import Optional

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Side, Alignment

# ---------------------------------------------------------------------------
# Config loading
# ---------------------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.json")

EXCEL_DATE_ORIGIN = datetime.date(1899, 12, 30)


def load_config() -> dict:
    if not os.path.exists(CONFIG_FILE):
        print(f"[错误] 找不到配置文件: {CONFIG_FILE}")
        sys.exit(1)
    with open(CONFIG_FILE, encoding="utf-8") as f:
        cfg = json.load(f)
    required = ("order_plan_file", "price_list_file", "factories")
    for key in required:
        if key not in cfg:
            print(f"[错误] config.json 缺少字段: {key}")
            sys.exit(1)
    return cfg


# ---------------------------------------------------------------------------
# Factory helpers
# ---------------------------------------------------------------------------


def get_year_folder(base_folder: str, year: int) -> str:
    """Return the contract folder for the given year, creating the full path if needed."""
    year_short = str(year)[-2:]
    folder = os.path.join(base_folder, f"{year_short}年")
    os.makedirs(folder, exist_ok=True)
    return folder


def get_template_file(cfg: dict, year: int) -> Optional[str]:
    """Find the latest contract in the current year's folder to use as template.
    Falls back to the previous year if no contracts exist yet for this year."""
    base_folder = cfg["base_folder"]
    prefix = cfg["file_prefix"]
    pattern = re.compile(rf"^{re.escape(prefix)}\d+\.xlsx$")

    for search_year in (year, year - 1):
        folder = get_year_folder(base_folder, search_year)
        candidates = [f for f in os.listdir(folder) if pattern.match(f)]
        if candidates:
            latest = sorted(candidates)[-1]
            return os.path.join(folder, latest)

    return None


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------


def excel_serial_to_date(serial) -> Optional[datetime.date]:
    if serial is None:
        return None
    try:
        n = int(serial)
    except (TypeError, ValueError):
        return None
    if n <= 0:
        return None
    return EXCEL_DATE_ORIGIN + datetime.timedelta(days=n)


def normalize_date(value) -> Optional[datetime.date]:
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.date() if isinstance(value, datetime.datetime) else value
    if isinstance(value, (int, float)):
        return excel_serial_to_date(value)
    return None


# ---------------------------------------------------------------------------
# Price list loading
# ---------------------------------------------------------------------------


def load_price_list(path: str) -> dict:
    """
    Returns a dict: sku -> {"出厂价": float|None, "含税价": float|None, "工厂": str,
                            "产品名称": str, "参数": str}
    Columns: SKU | 工厂 | 出厂价 | 含税价 | 产品名称(optional) | 参数(optional)
    Price list SKUs may be combined like "N03SR018BB/BR /WW"; we expand them.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    prices: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        sku_raw = str(row[0]).strip()
        factory = str(row[1]).strip() if row[1] else ""
        factory_price = float(row[2]) if row[2] is not None else None
        tax_price = float(row[3]) if row[3] is not None else None
        product_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        param = str(row[5]).strip() if len(row) > 5 and row[5] else ""

        for sku in _expand_sku_variants(sku_raw):
            prices[sku] = {
                "出厂价": factory_price,
                "含税价": tax_price,
                "工厂": factory,
                "产品名称": product_name,
                "参数": param,
            }

    return prices


def _expand_sku_variants(sku_raw: str) -> list[str]:
    """
    'N03SR018BB/BR /WW' -> ['N03SR018BB', 'N03SR018BR', 'N03SR018WW']
    'N02LH003NB'        -> ['N02LH003NB']
    """
    cleaned = sku_raw.replace(" ", "")
    parts = cleaned.split("/")
    if len(parts) == 1:
        return [parts[0]]
    first = parts[0]
    prefix = first[:-2]
    variants = [first] + [prefix + s for s in parts[1:] if s]
    return variants


# ---------------------------------------------------------------------------
# SKU detail + image lookup (scan existing contracts)
# ---------------------------------------------------------------------------


def build_sku_image_lookup(scan_dirs: list[str]) -> dict:
    """
    Scan all contracts for product images.
    Returns dict: sku -> image bytes (PNG or JPEG)
    Uses the anchor row in drawing1.xml to map image → SKU.
    """
    lookup: dict[str, bytes] = {}

    for scan_dir in scan_dirs:
        if not os.path.isdir(scan_dir):
            continue
        for root, _dirs, files in os.walk(scan_dir):
            for fname in files:
                if not fname.endswith(".xlsx") or fname.startswith("~"):
                    continue
                fpath = os.path.join(root, fname)
                try:
                    _extract_images_from_contract(fpath, lookup)
                except Exception:
                    pass

    return lookup


def _extract_images_from_contract(path: str, lookup: dict):
    """Extract SKU→image mapping from a single contract file."""
    with zipfile.ZipFile(path, "r") as z:
        names = z.namelist()
        drawing_path = "xl/drawings/drawing1.xml"
        rels_path = "xl/drawings/_rels/drawing1.xml.rels"
        if drawing_path not in names or rels_path not in names:
            return

        # rId → image file path
        rels_xml = z.read(rels_path).decode("utf-8")
        rid_to_img: dict[str, str] = {}
        for m in re.finditer(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml):
            target = m.group(2).replace("../", "xl/")
            rid_to_img[m.group(1)] = target

        # anchor row (0-indexed) → rId
        drawing_xml = z.read(drawing_path).decode("utf-8")
        anchors = re.findall(r"<xdr:twoCellAnchor.*?</xdr:twoCellAnchor>", drawing_xml, re.DOTALL)
        row_to_rid: dict[int, str] = {}
        for anchor in anchors:
            row_m = re.search(r"<xdr:from>.*?<xdr:row>(\d+)</xdr:row>", anchor, re.DOTALL)
            rid_m = re.search(r'r:embed="(rId\d+)"', anchor)
            if row_m and rid_m:
                row_to_rid[int(row_m.group(1))] = rid_m.group(1)

    # Read SKUs from the sheet (need a separate open outside the zip context)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return
    if "合同标的" not in wb.sheetnames:
        return
    ws = wb["合同标的"]
    sku_by_xl_row: dict[int, str] = {}
    for i, row in enumerate(ws.iter_rows(min_row=6, values_only=True)):
        sku = row[0]
        if not sku:
            continue
        sku = str(sku).strip()
        if sku and sku[0].isalnum():
            sku_by_xl_row[6 + i] = sku  # Excel row (1-indexed)

    # Map SKU → image bytes (drawing rows are 0-indexed, so +1 = Excel row)
    with zipfile.ZipFile(path, "r") as z:
        for draw_row, rid in row_to_rid.items():
            xl_row = draw_row + 1
            sku = sku_by_xl_row.get(xl_row)
            if sku and sku not in lookup:
                img_path = rid_to_img.get(rid, "")
                if img_path in z.namelist():
                    lookup[sku] = z.read(img_path)


def build_sku_detail_lookup(scan_dirs: list[str]) -> dict:
    """
    Scan all .xlsx files under scan_dirs for sheets named '合同标的'.
    Returns dict: sku -> {"产品名称": str, "参数": str}
    """
    lookup: dict[str, dict] = {}

    for scan_dir in scan_dirs:
        if not os.path.isdir(scan_dir):
            continue
        for root, _dirs, files in os.walk(scan_dir):
            for fname in files:
                if not fname.endswith(".xlsx") or fname.startswith("~"):
                    continue
                fpath = os.path.join(root, fname)
                try:
                    wb = openpyxl.load_workbook(fpath, data_only=True)
                except Exception as e:
                    print(f"  [警告] 无法读取 {fpath}: {e}")
                    continue
                if "合同标的" not in wb.sheetnames:
                    continue
                _extract_sku_details(wb["合同标的"], lookup, fpath)

    return lookup


def _extract_sku_details(ws, lookup: dict, fpath: str):
    h = {ws.cell(5, c).value: c for c in range(1, 12) if ws.cell(5, c).value}

    if "FNSKU" in h:
        col_name, col_param = 3, 5
    else:
        col_name, col_param = 2, 4

    for row in ws.iter_rows(min_row=6, values_only=True):
        sku = row[0]
        if not sku:
            continue
        sku = str(sku).strip()
        if not sku or not sku[0].isalnum():
            continue

        name = str(row[col_name - 1]).strip() if row[col_name - 1] else ""
        param = str(row[col_param - 1]).strip() if row[col_param - 1] else ""

        if sku not in lookup or (name and not lookup[sku]["产品名称"]):
            lookup[sku] = {"产品名称": name, "参数": param}


# ---------------------------------------------------------------------------
# Contract number generation
# ---------------------------------------------------------------------------


def next_contract_number(factory_cfg: dict, year: int) -> tuple[str, str, int]:
    folder = get_year_folder(factory_cfg["base_folder"], year)
    prefix = factory_cfg["file_prefix"]
    contract_prefix = factory_cfg["contract_prefix"]
    year_short = str(year)[-2:]

    max_seq = 0
    pattern = re.compile(rf"^{re.escape(prefix)}(\d+)\.xlsx$")
    for fname in os.listdir(folder):
        m = pattern.match(fname)
        if m:
            try:
                max_seq = max(max_seq, int(m.group(1)[-3:]))
            except ValueError:
                pass

    next_seq = max_seq + 1
    contract_number = f"{contract_prefix}{year}{next_seq:03d}"
    filename = f"{prefix}{year_short}{next_seq:03d}.xlsx"
    return contract_number, filename, next_seq


# ---------------------------------------------------------------------------
# Order plan parsing
# ---------------------------------------------------------------------------


def parse_order_plan(path: str, target_date: datetime.date) -> dict:
    """
    Returns dict keyed by (factory_name, is_larzonic).
    Larzonic sheet orders are kept separate from other sheets even for the same factory,
    because they use factory price (出厂价) while others use tax-inclusive price (含税价).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    all_orders: dict[tuple, list] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_orders = _parse_sheet(ws, sheet_name, target_date)
        is_larzonic = sheet_name == "Larzonic"
        for factory, items in sheet_orders.items():
            key = (factory, is_larzonic)
            all_orders.setdefault(key, []).extend(items)

    return all_orders


def _parse_sheet(ws, sheet_name: str, target_date: datetime.date) -> dict:
    if sheet_name == "脏衣篮订货":
        return _parse_zang_yi_lan(ws, sheet_name, target_date)
    elif sheet_name in ("铁网鞋架", "铁线"):
        return _parse_tie_wang_or_xian(ws, sheet_name, target_date)
    elif sheet_name == "Larzonic":
        return _parse_larzonic(ws, sheet_name, target_date)
    else:
        print(f"  [警告] 未知sheet: {sheet_name}，跳过")
        return {}


def _parse_zang_yi_lan(ws, sheet_name: str, target_date: datetime.date) -> dict:
    orders: dict[str, list] = {}
    date_cols = []
    for c in range(13, ws.max_column + 1, 2):
        d = normalize_date(ws.cell(4, c).value)
        if d is not None:
            date_cols.append((c, d))

    target_cols = [c for c, d in date_cols if d == target_date]
    if not target_cols:
        return {}

    for row in ws.iter_rows(min_row=9, values_only=True):
        sku = row[1]
        factory = row[2]
        if not sku or not factory:
            continue
        sku = str(sku).strip()
        factory = str(factory).strip()
        if not sku or not sku[0].isalnum():
            continue
        for c in target_cols:
            qty_raw = row[c - 1]
            if qty_raw is None:
                continue
            try:
                qty = int(qty_raw)
            except (TypeError, ValueError):
                continue
            if qty <= 0:
                continue
            orders.setdefault(factory, []).append({
                "sku": sku, "qty": qty, "chinese_name": "", "source_sheet": sheet_name,
            })

    return _merge_duplicate_skus(orders)


def _parse_tie_wang_or_xian(ws, sheet_name: str, target_date: datetime.date) -> dict:
    orders: dict[str, list] = {}
    date_cols = []
    for c in range(10, ws.max_column + 1, 2):
        d = normalize_date(ws.cell(5, c).value)
        if d is not None:
            date_cols.append((c, d))

    target_cols = [c for c, d in date_cols if d == target_date]
    if not target_cols:
        return {}

    for row in ws.iter_rows(min_row=10, values_only=True):
        sku = row[0]
        factory = row[1]
        chinese_name = row[2] or ""
        if not sku or not factory:
            continue
        sku = str(sku).strip()
        factory = str(factory).strip()
        chinese_name = str(chinese_name).strip()
        if not sku or not sku[0].isalnum():
            continue
        for c in target_cols:
            qty_raw = row[c - 1]
            if qty_raw is None:
                continue
            try:
                qty = int(qty_raw)
            except (TypeError, ValueError):
                continue
            if qty <= 0:
                continue
            orders.setdefault(factory, []).append({
                "sku": sku, "qty": qty, "chinese_name": chinese_name, "source_sheet": sheet_name,
            })

    return _merge_duplicate_skus(orders)


def _parse_larzonic(ws, sheet_name: str, target_date: datetime.date) -> dict:
    orders: dict[str, list] = {}
    date_cols = []
    for c in range(11, ws.max_column + 1):
        d = normalize_date(ws.cell(3, c).value)
        if d is not None:
            date_cols.append((c, d))

    target_cols = [c for c, d in date_cols if d == target_date]
    if not target_cols:
        return {}

    for row in ws.iter_rows(min_row=7, values_only=True):
        sku = row[0]
        factory = row[1]
        chinese_name = row[2] or ""
        if not sku or not factory:
            continue
        sku = str(sku).strip()
        factory = str(factory).strip()
        chinese_name = str(chinese_name).strip()
        if not sku or not sku[0].isalnum():
            continue
        for c in target_cols:
            qty_raw = row[c - 1]
            if qty_raw is None:
                continue
            try:
                qty = int(qty_raw)
            except (TypeError, ValueError):
                continue
            if qty <= 0:
                continue
            orders.setdefault(factory, []).append({
                "sku": sku, "qty": qty, "chinese_name": chinese_name,
                "source_sheet": sheet_name, "larzonic": True,
            })

    return _merge_duplicate_skus(orders)


def _merge_duplicate_skus(orders: dict) -> dict:
    merged: dict[str, list] = {}
    for factory, items in orders.items():
        seen: dict[str, dict] = {}
        for item in items:
            sku = item["sku"]
            if sku in seen:
                seen[sku]["qty"] += item["qty"]
            else:
                seen[sku] = dict(item)
        merged[factory] = list(seen.values())
    return merged


# ---------------------------------------------------------------------------
# Contract generation
# ---------------------------------------------------------------------------


def generate_contract(
    factory_name: str,
    factory_cfg: dict,
    orders: list,
    target_date: datetime.date,
    price_lookup: dict,
    sku_detail_lookup: dict,
    sku_image_lookup: dict,
    today: datetime.date,
    template_override: Optional[str] = None,
) -> Optional[str]:
    year = today.year
    contract_number, filename, seq = next_contract_number(factory_cfg, year)
    folder = get_year_folder(factory_cfg["base_folder"], year)
    out_path = os.path.join(folder, filename)

    template_path = template_override if template_override is not None else get_template_file(factory_cfg, year)
    if template_path is None:
        print(f"  [错误] 找不到 {factory_name} 的合同模板，跳过")
        return None

    orders_sorted = sorted(orders, key=lambda x: x["sku"])

    enriched = []
    for item in orders_sorted:
        sku = item["sku"]
        is_larzonic = item.get("larzonic", False)

        price_info = price_lookup.get(sku)
        if price_info is None:
            print(f"  [警告] 价格表中未找到SKU: {sku}，跳过")
            continue

        price_col_key = "出厂价" if is_larzonic else factory_cfg.get("price_col", "含税价")
        price = price_info.get(price_col_key)
        if price is None:
            print(f"  [警告] SKU {sku} 的价格为空，跳过")
            continue

        # 优先用历史合同详情，其次用价格表备注列，最后用订货计划里的中文名
        detail = sku_detail_lookup.get(sku)
        if detail and detail.get("产品名称"):
            product_name = detail["产品名称"]
            param = detail["参数"]
        elif price_info.get("产品名称"):
            product_name = price_info["产品名称"]
            param = price_info.get("参数", "")
        else:
            product_name = item.get("chinese_name", "")
            param = ""
            if not product_name:
                print(f"  [提示] SKU {sku} 无产品名称，建议在价格表补充")

        enriched.append({
            "sku": sku,
            "product_name": product_name,
            "param": param,
            "qty": item["qty"],
            "price": price,
        })

    if not enriched:
        print(f"  [跳过] {factory_name}: 没有有效SKU")
        return None

    is_larzonic = any(item.get("larzonic") for item in orders)

    shutil.copy2(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)
    _fill_purchase_sheet(wb["采购单"], contract_number, today)
    _fill_contract_sheet(wb["合同标的"], contract_number, target_date, enriched, is_larzonic=is_larzonic)
    wb.save(out_path)
    _rebuild_contract_drawings(template_path, out_path, enriched, sku_image_lookup)

    total_qty = sum(item["qty"] for item in enriched)
    total_amount = sum(item["qty"] * item["price"] for item in enriched)
    print(
        f"生成合同：{filename}  ({factory_name}, {len(enriched)}个SKU, "
        f"合计{total_qty}套, ¥{total_amount:,.0f})"
    )
    return filename


def _fill_purchase_sheet(ws, contract_number: str, today: datetime.date):
    ws.cell(2, 2).value = contract_number
    new_date_str = f"{today.year}.{today.month}.{today.day}"
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and "签订日期" in str(cell.value):
                cell.value = re.sub(
                    r"(签订日期[：:])[\d./]+",
                    rf"\g<1>{new_date_str}",
                    str(cell.value),
                )


ROW_HEIGHT_PT = 138  # fixed data-row height — keeps printed images the same size every time


_THIN = Side(style="thin")
_FULL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NO_BORDER = Border()

# Per-column alignment for data rows (col index 1-based)
_COL_ALIGN = {
    1: Alignment(wrap_text=True, vertical="center", horizontal="center"),
    2: Alignment(wrap_text=True, vertical="center", horizontal="center"),
    3: Alignment(wrap_text=False, vertical="center", horizontal="center"),
    4: Alignment(wrap_text=True, vertical="top",    horizontal="left"),
    6: Alignment(wrap_text=True, vertical="center", horizontal="center"),
    7: Alignment(wrap_text=True, vertical="center", horizontal="center"),
    8: Alignment(wrap_text=True, vertical="center", horizontal="center"),
}


def _fill_contract_sheet(ws, contract_number: str, target_date: datetime.date, enriched: list, is_larzonic: bool = False):
    ws.cell(1, 2).value = contract_number
    ws.cell(2, 2).value = datetime.datetime(target_date.year, target_date.month, target_date.day)

    # Fix price column header: template always says "出厂价"; non-Larzonic contracts use 含税价
    if not is_larzonic:
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(5, c)
            if cell.value and "出厂价" in str(cell.value):
                cell.value = str(cell.value).replace("出厂价", "含税价")

    data_area_merges = [mr for mr in list(ws.merged_cells.ranges) if mr.min_row >= 6]
    for mr in data_area_merges:
        ws.unmerge_cells(str(mr))

    # Capture font from template row 6 (宋体 11pt) for reuse
    tmpl_font = copy(ws.cell(6, 1).font)

    _clear_data_rows(ws, start_row=6)

    data_start_row = 6
    for i, item in enumerate(enriched):
        r = data_start_row + i
        ws.row_dimensions[r].height = ROW_HEIGHT_PT
        ws.cell(r, 1).value = item["sku"]
        ws.cell(r, 2).value = item["product_name"]
        ws.cell(r, 4).value = item["param"]
        ws.merge_cells(f"D{r}:E{r}")
        ws.cell(r, 6).value = item["qty"]
        ws.cell(r, 7).value = item["price"]
        ws.cell(r, 8).value = f"=F{r}*G{r}"

        for c in range(1, 9):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.font = copy(tmpl_font)
            cell.alignment = copy(_COL_ALIGN.get(c, _COL_ALIGN[1]))
            cell.border = _FULL_BORDER

    last_data_row = data_start_row + len(enriched) - 1
    total_row = last_data_row + 1
    ws.cell(total_row, 1).value = "合计"
    ws.cell(total_row, 6).value = f"=SUM(F{data_start_row}:F{last_data_row})"
    ws.cell(total_row, 8).value = f"=SUM(H{data_start_row}:H{last_data_row})"

    # Remove borders and reset row height for all template rows below the total row
    for r in range(total_row + 1, ws.max_row + 1):
        ws.row_dimensions[r].height = None
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.border = _NO_BORDER


def _clear_data_rows(ws, start_row: int):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def _rebuild_contract_drawings(
    template_path: str,
    out_path: str,
    enriched: list,
    sku_image_lookup: dict,
):
    """
    Rebuild the contract zip with correct product images in 合同标的 (drawing1.xml).
    Structural drawings on other sheets (logos, signatures) are copied from the template.
    """
    DATA_START_ROW = 6  # first data row in 合同标的 (1-indexed Excel)

    # Build content-hash → (rId, media_path, bytes) for deduplication
    seen_hashes: dict[int, str] = {}   # content hash → rId
    rid_to_path: dict[str, str] = {}   # rId → xl/media/... path
    new_media: dict[str, bytes] = {}   # xl/media/... → bytes
    sku_to_rid: dict[str, str] = {}    # sku → rId

    rid_counter = 1
    skus_with_images = [
        (i, item["sku"]) for i, item in enumerate(enriched)
        if item["sku"] in sku_image_lookup
    ]

    for _, sku in skus_with_images:
        img_bytes = sku_image_lookup[sku]
        h = hash(img_bytes)
        if h not in seen_hashes:
            rid = f"rId{rid_counter}"
            ext = "jpeg" if img_bytes[:3] == b"\xff\xd8\xff" else "png"
            media_path = f"xl/media/prod_image{rid_counter}.{ext}"
            seen_hashes[h] = rid
            rid_to_path[rid] = media_path
            new_media[media_path] = img_bytes
            rid_counter += 1
        sku_to_rid[sku] = seen_hashes[h]

    # Build new drawing1.xml
    # Col offsets from existing contracts (col C ≈ 30mm wide)
    # Row offsets derived from ROW_HEIGHT_PT so images are always the same printed size
    COL, COL_OFF_FROM, COL_OFF_TO = 2, 34290, 1129030
    _row_padding = 95000  # ~2.6mm top/bottom padding
    ROW_OFF_FROM = _row_padding
    ROW_OFF_TO = ROW_HEIGHT_PT * 12700 - _row_padding  # 1657600 EMU for 138pt rows

    anchor_parts = []
    for pic_id, (i, sku) in enumerate(skus_with_images, start=2):
        draw_row = DATA_START_ROW + i - 1   # 0-indexed drawing row
        rid = sku_to_rid[sku]
        anchor_parts.append(
            f'<xdr:twoCellAnchor editAs="oneCell">'
            f"<xdr:from><xdr:col>{COL}</xdr:col><xdr:colOff>{COL_OFF_FROM}</xdr:colOff>"
            f"<xdr:row>{draw_row}</xdr:row><xdr:rowOff>{ROW_OFF_FROM}</xdr:rowOff></xdr:from>"
            f"<xdr:to><xdr:col>{COL}</xdr:col><xdr:colOff>{COL_OFF_TO}</xdr:colOff>"
            f"<xdr:row>{draw_row}</xdr:row><xdr:rowOff>{ROW_OFF_TO}</xdr:rowOff></xdr:to>"
            f'<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="{pic_id}" name="产品图{pic_id}"/>'
            f'<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr></xdr:nvPicPr>'
            f'<xdr:blipFill><a:blip r:embed="{rid}"/>'
            f"<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>"
            f'<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="962025" cy="1270000"/></a:xfrm>'
            f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln w=\"9525\"><a:noFill/></a:ln></xdr:spPr></xdr:pic>"
            f"<xdr:clientData/></xdr:twoCellAnchor>"
        )

    new_drawing_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        + "".join(anchor_parts)
        + "</xdr:wsDr>"
    ).encode("utf-8")

    new_drawing_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + "".join(
            f'<Relationship Id="{rid}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
            f'Target="{path.replace("xl/", "../")}"/>'
            for rid, path in rid_to_path.items()
        )
        + "</Relationships>"
    ).encode("utf-8")

    # Read structural pieces from template (non-product drawings and rels)
    with zipfile.ZipFile(template_path, "r") as tmpl_zip:
        tmpl_structural: dict[str, bytes] = {}
        for name in tmpl_zip.namelist():
            if name.startswith("xl/drawings/") and name != "xl/drawings/drawing1.xml" \
                    and name != "xl/drawings/_rels/drawing1.xml.rels":
                tmpl_structural[name] = tmpl_zip.read(name)

        sheet_drawing_ref: dict[str, str] = {}
        for name in tmpl_zip.namelist():
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                content = tmpl_zip.read(name).decode("utf-8")
                m = re.search(r"<drawing[^/]*/>", content)
                if m:
                    sheet_drawing_ref[os.path.basename(name)] = m.group(0)

        sheet_rels = {
            name: tmpl_zip.read(name)
            for name in tmpl_zip.namelist()
            if name.startswith("xl/worksheets/_rels/")
        }

    # Rewrite the output zip
    tmp_path = out_path + ".tmp"
    skip = (
        set(tmpl_structural) | set(sheet_rels) | set(new_media)
        | {"xl/drawings/drawing1.xml", "xl/drawings/_rels/drawing1.xml.rels"}
    )

    with zipfile.ZipFile(out_path, "r") as out_zip, \
         zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as new_zip:

        for item in out_zip.infolist():
            if item.filename in skip:
                continue
            content = out_zip.read(item.filename)
            fname = os.path.basename(item.filename)
            if fname in sheet_drawing_ref and item.filename.startswith("xl/worksheets/"):
                text = content.decode("utf-8")
                if "<drawing" not in text:
                    tag = sheet_drawing_ref[fname]
                    r_ns = 'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
                    if r_ns not in text:
                        text = text.replace("<worksheet ", f"<worksheet {r_ns} ", 1)
                    text = text.replace("</worksheet>", f"{tag}</worksheet>")
                    content = text.encode("utf-8")
            new_zip.writestr(item, content)

        for name, data in tmpl_structural.items():
            new_zip.writestr(name, data)
        for name, data in sheet_rels.items():
            new_zip.writestr(name, data)
        for name, data in new_media.items():
            new_zip.writestr(name, data)

        new_zip.writestr("xl/drawings/drawing1.xml", new_drawing_xml)
        new_zip.writestr("xl/drawings/_rels/drawing1.xml.rels", new_drawing_rels)

    os.replace(tmp_path, out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    if len(sys.argv) < 2:
        print("用法: python generate_contract.py <交货日期>")
        print("示例: python generate_contract.py 2026-06-01")
        sys.exit(1)

    date_str = sys.argv[1]
    try:
        target_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        print(f"日期格式错误: {date_str}，请使用 YYYY-MM-DD 格式")
        sys.exit(1)

    config = load_config()
    factory_configs = config["factories"]
    order_plan_file = config["order_plan_file"]
    price_list_file = config["price_list_file"]

    # Derive scan dirs from factory base_folders
    scan_dirs = [cfg["base_folder"] for cfg in factory_configs.values()]

    today = datetime.date.today()
    print(f"目标交货日期: {target_date}  |  今日: {today}")
    print()

    print("正在加载价格表…")
    price_lookup = load_price_list(price_list_file)
    print(f"  加载了 {len(price_lookup)} 个SKU的价格")

    print("正在扫描合同文件以建立SKU详情库和图片库…")
    sku_detail_lookup = build_sku_detail_lookup(scan_dirs)
    sku_image_lookup = build_sku_image_lookup(scan_dirs)
    print(f"  扫描到 {len(sku_detail_lookup)} 个SKU的详情，{len(sku_image_lookup)} 个SKU的图片")

    print("正在解析订货计划…")
    orders_by_factory = parse_order_plan(order_plan_file, target_date)

    if not orders_by_factory:
        print(f"\n未找到 {target_date} 的订单，退出。")
        sys.exit(0)

    print(f"  找到 {len(orders_by_factory)} 个工厂的订单:")
    for (factory, is_larzonic), items in orders_by_factory.items():
        tag = "（Larzonic 出厂价）" if is_larzonic else "（含税价）"
        print(f"    {factory}{tag}: {len(items)} 个SKU")
    print()

    # Snapshot template paths before any contracts are generated, so newly created contracts
    # don't accidentally become the template for subsequent contracts in the same run.
    template_snapshot: dict[str, Optional[str]] = {
        name: get_template_file(cfg, today.year)
        for name, cfg in factory_configs.items()
    }

    generated = []
    for (factory_name, is_larzonic) in sorted(orders_by_factory.keys()):
        if factory_name not in factory_configs:
            print(f"  [警告] 未配置工厂: {factory_name}，请在 config.json 的 factories 里添加，跳过")
            continue
        tag = "（Larzonic 出厂价）" if is_larzonic else "（含税价）"
        print(f"\n正在生成 {factory_name}{tag} 合同…")
        result = generate_contract(
            factory_name=factory_name,
            factory_cfg=factory_configs[factory_name],
            orders=orders_by_factory[(factory_name, is_larzonic)],
            target_date=target_date,
            price_lookup=price_lookup,
            sku_detail_lookup=sku_detail_lookup,
            sku_image_lookup=sku_image_lookup,
            today=today,
            template_override=template_snapshot.get(factory_name),
        )
        if result:
            generated.append(result)

    if not generated:
        print("\n没有生成任何合同。")
    else:
        print(f"\n共生成 {len(generated)} 份合同。")


if __name__ == "__main__":
    main()
